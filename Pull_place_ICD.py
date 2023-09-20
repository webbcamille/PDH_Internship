import openpyxl
import os
import re

# Get user input for output file name and save the Excel sheet
output_name = input("Name of output file: ").strip()

# Check if the user included an extension. If not, append '.xlsx' to the filename
if not output_name.endswith('.xlsx'):
    output_name += '.xlsx'

path = r"C:\Users\Administrator\Documents\ParaDocsHealthPrototypes-main\ParaDocsHealthPrototypes-main\Output Files"

# Create the full path by joining the directory path and filename
full_path = os.path.join(path, output_name)


def find_end_delimiter(file_name):
    '''Find the end delimiter dynamically by reading the file'''
    with open(file_name, 'r') as file:
        lines = file.readlines()

    # Identify the start of the desired section
    try:
        start_index = lines.index('- Assessment -\n') + 1

        # Find the end index by searching for the second empty line after '- Assessment -'
        empty_line_count = 0
        for i, line in enumerate(lines[start_index:], start=start_index):
            if line.strip() == '':
                empty_line_count += 1
                if empty_line_count == 2:
                    end_index = i
                    break

        if empty_line_count >= 2:
            # If two empty lines found after '- Assessment -', consider the line after that as the end delimiter
            return end_index
        else:
            # If not enough empty lines found after '- Assessment -', consider the last line as the end delimiter
            return len(lines) - 1
    except ValueError:
        print("Section not found")
        return -1

def extract_ICD(file_name, end_index):
    '''Extracts the section between - Assessment - and - Plan -'''
    with open(file_name, 'r') as file:
        lines = file.readlines()

    # Identify the start and end of the desired section
    try:
        start_index = lines.index('- Assessment -\n') + 1
        return lines[start_index:end_index]
    
    except ValueError:
        print("Section not found")
        return []


def remove_empty_rows(full_path):
    wb = openpyxl.load_workbook(full_path)
    ws = wb.active

    rows_to_delete = [
        row[0].row
        for row in ws.iter_rows()
        if all(cell.value is None for cell in row)
    ]
    # Delete rows from bottom to top (to avoid reindexing problems)
    for row_index in reversed(rows_to_delete):
        ws.delete_rows(row_index)

    wb.save(full_path)


def main():
    # Get user input for file name
    file_name = input("Enter the path to the .txt file: ")
    
    end_index = find_end_delimiter(file_name)

    try:
        end_index = int(end_index)
    except ValueError:
        print("Invalid end index. Please make sure 'end_index' is an integer.")
        return

    # Extract section from text file
    ICD_data = extract_ICD(file_name, end_index)
    
    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Insert each ICD into the Excel sheet
    for index, line in enumerate(ICD_data, 1):
        
        if not line.strip() or ' ' not in line:
            continue
    
        # Split each line into code and description
        code, description = line.split(None, 1)  # Split at the first occurrence of whitespace
        ws.cell(row=index, column=1, value=code.strip())
        ws.cell(row=index, column=2, value=description.strip())
    

    # Save the workbook to the specified path
    wb.save(full_path)
    
    # Call the function with the path to your Excel file
    remove_empty_rows(full_path)
    
if __name__ == '__main__':
    main()
